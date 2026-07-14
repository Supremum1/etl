from __future__ import annotations

import csv
from collections.abc import Iterator
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def scalar_to_text(value: Any) -> str | None:
    if value is None:
        return None
    if hasattr(value, "isoformat"):
        value = value.isoformat()
    return str(value).replace("\x00", "")


def detect_file(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return "csv"
    if suffix in {".xlsx", ".xlsm"}:
        return "xlsx"
    raise ValueError(f"Unsupported input format '{suffix}'. Use CSV or XLSX.")


def find_default_file(dataset: str, data_dir: Path) -> Path:
    for suffix in (".csv", ".xlsx", ".xlsm"):
        candidate = data_dir / f"{dataset}{suffix}"
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        f"Input file not found. Put {dataset}.csv or {dataset}.xlsx into {data_dir}, "
        "or pass --file explicitly."
    )


def iter_csv(path: Path, delimiter: str | None = None) -> tuple[list[str], Iterator[list[str | None]]]:
    handle = path.open("r", newline="", encoding="utf-8-sig")
    sample = handle.read(65536)
    handle.seek(0)
    if delimiter is None:
        dialect = csv.Sniffer().sniff(sample) if sample else csv.excel
        reader = csv.reader(handle, dialect)
    else:
        reader = csv.reader(handle, delimiter=delimiter)
    headers = next(reader)

    def rows() -> Iterator[list[str | None]]:
        try:
            for row in reader:
                yield [scalar_to_text(cell) if cell != "" else None for cell in row]
        finally:
            handle.close()

    return headers, rows()


def iter_xlsx(path: Path, sheet: str | None = None) -> tuple[list[str], Iterator[list[str | None]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook[workbook.sheetnames[0]]
    row_iter = worksheet.iter_rows(values_only=True)
    headers = [scalar_to_text(value) or "" for value in next(row_iter)]

    def rows() -> Iterator[list[str | None]]:
        try:
            for row in row_iter:
                yield [scalar_to_text(value) for value in row]
        finally:
            workbook.close()

    return headers, rows()


def iter_input(
    path: Path,
    sheet: str | None = None,
    delimiter: str | None = None,
) -> tuple[str, list[str], Iterator[list[str | None]]]:
    fmt = detect_file(path)
    if fmt == "csv":
        headers, rows = iter_csv(path, delimiter)
    else:
        headers, rows = iter_xlsx(path, sheet)
    return fmt, headers, rows
